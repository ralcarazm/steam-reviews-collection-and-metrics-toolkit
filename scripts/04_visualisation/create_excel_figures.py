#!/usr/bin/env python3
"""
create_excel_figures.py

Create one Excel workbook per figure from previously generated CSV tables.

This script:
- reads a per-game JSON configuration file;
- loads selected CSV tables produced by earlier analysis scripts;
- creates one Excel file per figure in results/<game_slug>/figures/;
- writes the exact source table used for each figure into a 'data' sheet;
- creates an Excel chart in a 'chart' sheet;
- writes figure metadata into a 'metadata' sheet;
- writes a figure index CSV, JSON summary, and log.

Usage example:

    python scripts/04_visualisation/create_excel_figures.py \
        --config config/games/example_game.json

Author:
    Rubén Alcaraz Martínez

Licence:
    GNU General Public License v3.0
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font


# ============================================================================
# DATA MODELS
# ============================================================================

@dataclass
class GameConfig:
    """Configuration values for one figure-generation run."""
    app_id: int
    game_slug: str
    game_title: str


@dataclass
class FigureSpec:
    """Specification for one Excel figure."""
    figure_id: str
    filename: str
    family: str
    title: str
    chart_type: str
    source_table: str
    transform_name: str
    category_column: str
    value_columns: List[str]
    notes: str = ""


# ============================================================================
# ARGUMENT PARSING
# ============================================================================

def parse_arguments() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Create Excel figure workbooks from existing metrics tables."
    )
    parser.add_argument(
        "--config",
        required=True,
        help="Path to the JSON configuration file for the target game.",
    )
    return parser.parse_args()


# ============================================================================
# CONFIGURATION LOADING
# ============================================================================

def load_game_config(config_path: Path) -> GameConfig:
    """Load and validate the game configuration file."""
    if not config_path.exists():
        raise FileNotFoundError(f"Configuration file not found: {config_path}")

    with config_path.open("r", encoding="utf-8") as handle:
        raw_config = json.load(handle)

    required_fields = ["app_id", "game_slug", "game_title"]
    missing_fields = [field for field in required_fields if field not in raw_config]
    if missing_fields:
        raise ValueError(
            f"Missing required configuration fields: {', '.join(missing_fields)}"
        )

    return GameConfig(
        app_id=int(raw_config["app_id"]),
        game_slug=str(raw_config["game_slug"]).strip(),
        game_title=str(raw_config["game_title"]).strip(),
    )


# ============================================================================
# PATH MANAGEMENT
# ============================================================================

def get_repository_root() -> Path:
    """
    Resolve the repository root assuming this file lives at:
    scripts/04_visualisation/create_excel_figures.py
    """
    return Path(__file__).resolve().parents[2]


def build_paths(repository_root: Path, config: GameConfig) -> Dict[str, Path]:
    """Build all required input and output paths."""
    results_root = repository_root / "results" / config.game_slug
    processed_root = repository_root / "data" / "processed" / "steam_reviews" / config.game_slug

    paths = {
        "results_root": results_root,
        "tables_root": results_root / "tables",
        "figures_root": results_root / "figures",
        "basic_figures_dir": results_root / "figures" / "basic_metrics",
        "temporal_figures_dir": results_root / "figures" / "temporal_metrics",
        "text_figures_dir": results_root / "figures" / "text_metrics",
        "emotion_figures_dir": results_root / "figures" / "emotion_metrics",
        "theme_figures_dir": results_root / "figures" / "theme_metrics",
        "metrics_dir": processed_root / "metrics",
    }

    paths["figure_index_csv"] = paths["figures_root"] / f"{config.game_slug}_figure_index.csv"
    paths["summary_json"] = paths["metrics_dir"] / f"{config.game_slug}_figures_summary.json"
    paths["analysis_log"] = paths["metrics_dir"] / f"{config.game_slug}_figures.log"

    return paths


def ensure_directories(paths: Dict[str, Path]) -> None:
    """Create required output directories."""
    for key in [
        "figures_root",
        "basic_figures_dir",
        "temporal_figures_dir",
        "text_figures_dir",
        "emotion_figures_dir",
        "theme_figures_dir",
        "metrics_dir",
    ]:
        paths[key].mkdir(parents=True, exist_ok=True)


# ============================================================================
# LOGGING
# ============================================================================

def configure_logging(log_file: Path) -> None:
    """Configure console and file logging."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )


# ============================================================================
# FIGURE SPECIFICATIONS
# ============================================================================

def get_figure_specs(config: GameConfig) -> List[FigureSpec]:
    """Return the default list of figure specifications."""
    game_slug = config.game_slug

    return [
        FigureSpec(
            figure_id="figure_01",
            filename="figure_01_polarity_distribution.xlsx",
            family="basic_metrics",
            title="Polarity distribution",
            chart_type="bar",
            source_table=f"tables/basic_metrics/{game_slug}_polarity_summary.csv",
            transform_name="polarity_distribution",
            category_column="polarity",
            value_columns=["review_count"],
            notes="Based on positive and negative review totals.",
        ),
        FigureSpec(
            figure_id="figure_02",
            filename="figure_02_language_distribution_top10.xlsx",
            family="basic_metrics",
            title="Language distribution (top 10)",
            chart_type="bar",
            source_table=f"tables/basic_metrics/{game_slug}_language_distribution.csv",
            transform_name="language_distribution_top10",
            category_column="language",
            value_columns=["review_count"],
            notes="Top 10 languages ranked by review count.",
        ),
        FigureSpec(
            figure_id="figure_03",
            filename="figure_03_polarity_by_playtime_band.xlsx",
            family="basic_metrics",
            title="Polarity by playtime band",
            chart_type="bar_stacked",
            source_table=f"tables/basic_metrics/{game_slug}_cross_tab_polarity_by_playtime_band.csv",
            transform_name="polarity_by_playtime_band",
            category_column="playtime_at_review_band",
            value_columns=["negative_reviews", "positive_reviews"],
            notes="Counts of positive and negative reviews by playtime band.",
        ),
        FigureSpec(
            figure_id="figure_04",
            filename="figure_04_monthly_review_volume.xlsx",
            family="temporal_metrics",
            title="Monthly review volume",
            chart_type="line",
            source_table=f"tables/temporal_metrics/{game_slug}_monthly_review_volume_extended.csv",
            transform_name="monthly_review_volume",
            category_column="review_created_date_month",
            value_columns=["review_count"],
            notes="Review count by month.",
        ),
        FigureSpec(
            figure_id="figure_05",
            filename="figure_05_monthly_polarity_trends.xlsx",
            family="temporal_metrics",
            title="Monthly polarity trends",
            chart_type="line",
            source_table=f"tables/temporal_metrics/{game_slug}_monthly_polarity_trends.csv",
            transform_name="monthly_polarity_trends",
            category_column="review_created_date_month",
            value_columns=["positive_reviews", "negative_reviews"],
            notes="Positive and negative review counts by month.",
        ),
        FigureSpec(
            figure_id="figure_06",
            filename="figure_06_monthly_positive_percentage.xlsx",
            family="temporal_metrics",
            title="Monthly positive percentage",
            chart_type="line",
            source_table=f"tables/temporal_metrics/{game_slug}_monthly_polarity_trends.csv",
            transform_name="monthly_positive_percentage",
            category_column="review_created_date_month",
            value_columns=["positive_percentage"],
            notes="Percentage of positive reviews by month.",
        ),
        FigureSpec(
            figure_id="figure_07",
            filename="figure_07_monthly_language_trends_top5.xlsx",
            family="temporal_metrics",
            title="Monthly language trends (top 5)",
            chart_type="line",
            source_table=f"tables/temporal_metrics/{game_slug}_monthly_language_trends.csv",
            transform_name="monthly_language_trends_top5",
            category_column="review_created_date_month",
            value_columns=[],
            notes="Top 5 languages by global review count across months.",
        ),
        FigureSpec(
            figure_id="figure_08",
            filename="figure_08_developer_response_timeline.xlsx",
            family="temporal_metrics",
            title="Developer response timeline",
            chart_type="line",
            source_table=f"tables/temporal_metrics/{game_slug}_developer_response_timeline.csv",
            transform_name="developer_response_timeline",
            category_column="review_created_date_month",
            value_columns=["reviews_with_developer_response"],
            notes="Reviews with developer response by review-creation month.",
        ),
        FigureSpec(
            figure_id="figure_09",
            filename="figure_09_top_unigrams_positive.xlsx",
            family="text_metrics",
            title="Top unigrams in positive reviews",
            chart_type="bar",
            source_table="tables/text_metrics/global/top_unigrams_positive.csv",
            transform_name="top_unigrams_positive",
            category_column="token",
            value_columns=["frequency"],
            notes="Top 20 unigrams in positive reviews.",
        ),
        FigureSpec(
            figure_id="figure_10",
            filename="figure_10_top_unigrams_negative.xlsx",
            family="text_metrics",
            title="Top unigrams in negative reviews",
            chart_type="bar",
            source_table="tables/text_metrics/global/top_unigrams_negative.csv",
            transform_name="top_unigrams_negative",
            category_column="token",
            value_columns=["frequency"],
            notes="Top 20 unigrams in negative reviews.",
        ),
        FigureSpec(
            figure_id="figure_11",
            filename="figure_11_distinctive_terms_positive_vs_negative.xlsx",
            family="text_metrics",
            title="Distinctive terms: positive vs negative",
            chart_type="bar",
            source_table="tables/text_metrics/global/distinctive_terms_positive_vs_negative.csv",
            transform_name="distinctive_terms_positive_vs_negative",
            category_column="token",
            value_columns=["difference_per_1000"],
            notes="Top 20 terms most overrepresented in positive reviews.",
        ),
        FigureSpec(
            figure_id="figure_12",
            filename="figure_12_distinctive_terms_negative_vs_positive.xlsx",
            family="text_metrics",
            title="Distinctive terms: negative vs positive",
            chart_type="bar",
            source_table="tables/text_metrics/global/distinctive_terms_negative_vs_positive.csv",
            transform_name="distinctive_terms_negative_vs_positive",
            category_column="token",
            value_columns=["difference_per_1000"],
            notes="Top 20 terms most overrepresented in negative reviews.",
        ),
        FigureSpec(
            figure_id="figure_13",
            filename="figure_13_emotion_distribution_summary.xlsx",
            family="emotion_metrics",
            title="Emotion distribution summary",
            chart_type="bar",
            source_table=f"tables/emotion_metrics/{game_slug}_emotion_distribution_summary.csv",
            transform_name="emotion_distribution_summary",
            category_column="emotion",
            value_columns=["mean_per_100_tokens"],
            notes="Mean emotion signal per 100 tokens.",
        ),
        FigureSpec(
            figure_id="figure_14",
            filename="figure_14_emotion_by_polarity.xlsx",
            family="emotion_metrics",
            title="Emotion profiles by polarity",
            chart_type="bar",
            source_table=f"tables/emotion_metrics/{game_slug}_emotion_by_polarity.csv",
            transform_name="emotion_by_polarity",
            category_column="emotion",
            value_columns=["positive", "negative"],
            notes="Emotion mean per 100 tokens for positive and negative segments.",
        ),
        FigureSpec(
            figure_id="figure_15",
            filename="figure_15_emotion_by_month_selected.xlsx",
            family="emotion_metrics",
            title="Selected emotions by month",
            chart_type="line",
            source_table=f"tables/emotion_metrics/{game_slug}_emotion_by_month.csv",
            transform_name="emotion_by_month_selected",
            category_column="review_created_year_month",
            value_columns=[
                "joy_mean_per_100_tokens",
                "trust_mean_per_100_tokens",
                "negative_mean_per_100_tokens",
                "anger_mean_per_100_tokens",
            ],
            notes="Selected emotional dimensions over time.",
        ),
        FigureSpec(
            figure_id="figure_16",
            filename="figure_16_theme_distribution_summary.xlsx",
            family="theme_metrics",
            title="Theme distribution summary",
            chart_type="bar",
            source_table=f"tables/theme_metrics/{game_slug}_theme_distribution_summary.csv",
            transform_name="theme_distribution_summary",
            category_column="theme",
            value_columns=["review_share_with_theme"],
            notes="Percentage of analysed reviews containing each theme.",
        ),
        FigureSpec(
            figure_id="figure_17",
            filename="figure_17_theme_by_polarity.xlsx",
            family="theme_metrics",
            title="Themes by polarity",
            chart_type="bar",
            source_table=f"tables/theme_metrics/{game_slug}_theme_by_polarity.csv",
            transform_name="theme_by_polarity",
            category_column="theme",
            value_columns=["positive", "negative"],
            notes="Theme review share for positive and negative segments.",
        ),
        FigureSpec(
            figure_id="figure_18",
            filename="figure_18_theme_by_month_selected.xlsx",
            family="theme_metrics",
            title="Selected themes by month",
            chart_type="line",
            source_table=f"tables/theme_metrics/{game_slug}_theme_by_month.csv",
            transform_name="theme_by_month_selected",
            category_column="review_created_year_month",
            value_columns=[
                "difficulty_review_share",
                "combat_review_share",
                "performance_technical_review_share",
                "story_review_share",
            ],
            notes="Selected themes over time.",
        ),
    ]


# ============================================================================
# INPUT HANDLING
# ============================================================================

def resolve_source_table_path(paths: Dict[str, Path], source_table: str) -> Path:
    """Resolve a source table path relative to results/<game_slug>/."""
    return paths["results_root"] / source_table


def load_source_table(source_path: Path) -> pd.DataFrame:
    """Load one CSV table."""
    if not source_path.exists():
        raise FileNotFoundError(f"Source table not found: {source_path}")
    return pd.read_csv(source_path, encoding="utf-8-sig", low_memory=False)


# ============================================================================
# TRANSFORMS
# ============================================================================

def transform_polarity_distribution(df: pd.DataFrame) -> pd.DataFrame:
    """Transform polarity summary into chart-ready data."""
    metric_to_label = {
        "positive_reviews": "positive",
        "negative_reviews": "negative",
    }
    temp = df[df["metric"].isin(metric_to_label.keys())].copy()
    temp["polarity"] = temp["metric"].map(metric_to_label)
    temp["review_count"] = pd.to_numeric(temp["value"], errors="coerce")
    return temp[["polarity", "review_count"]].reset_index(drop=True)


def transform_language_distribution_top10(df: pd.DataFrame) -> pd.DataFrame:
    """Select top 10 languages by review count."""
    temp = df.copy()
    temp["review_count"] = pd.to_numeric(temp["review_count"], errors="coerce")
    temp = temp.sort_values(["review_count", "language"], ascending=[False, True]).head(10)
    return temp[["language", "review_count"]].reset_index(drop=True)


def transform_polarity_by_playtime_band(df: pd.DataFrame) -> pd.DataFrame:
    """Rename polarity columns for chart use."""
    temp = df.copy()
    rename_map = {}
    if "False" in temp.columns:
        rename_map["False"] = "negative_reviews"
    if "True" in temp.columns:
        rename_map["True"] = "positive_reviews"
    temp = temp.rename(columns=rename_map)
    return temp[["playtime_at_review_band", "negative_reviews", "positive_reviews"]].reset_index(drop=True)


def transform_monthly_review_volume(df: pd.DataFrame) -> pd.DataFrame:
    """Return month and review count."""
    return df[["review_created_date_month", "review_count"]].copy()


def transform_monthly_polarity_trends(df: pd.DataFrame) -> pd.DataFrame:
    """Return month and positive/negative review counts."""
    return df[["review_created_date_month", "positive_reviews", "negative_reviews"]].copy()


def transform_monthly_positive_percentage(df: pd.DataFrame) -> pd.DataFrame:
    """Return month and positive percentage."""
    return df[["review_created_date_month", "positive_percentage"]].copy()


def transform_monthly_language_trends_top5(df: pd.DataFrame) -> pd.DataFrame:
    """Pivot top 5 languages by global review count across months."""
    temp = df.copy()
    totals = (
        temp.groupby("language", dropna=False)["review_count"]
        .sum()
        .reset_index()
        .sort_values(["review_count", "language"], ascending=[False, True])
        .head(5)
    )
    selected_languages = totals["language"].astype(str).tolist()

    temp = temp[temp["language"].astype(str).isin(selected_languages)].copy()
    pivot = temp.pivot_table(
        index="review_created_date_month",
        columns="language",
        values="review_count",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    pivot.columns = ["review_created_date_month"] + [str(col) for col in pivot.columns[1:]]
    return pivot.sort_values("review_created_date_month").reset_index(drop=True)


def transform_developer_response_timeline(df: pd.DataFrame) -> pd.DataFrame:
    """Return month and developer response count."""
    return df[["review_created_date_month", "reviews_with_developer_response"]].copy()


def transform_top_unigrams_positive(df: pd.DataFrame) -> pd.DataFrame:
    """Select top 20 positive unigrams."""
    return df.head(20)[["token", "frequency"]].copy()


def transform_top_unigrams_negative(df: pd.DataFrame) -> pd.DataFrame:
    """Select top 20 negative unigrams."""
    return df.head(20)[["token", "frequency"]].copy()


def transform_distinctive_terms_positive_vs_negative(df: pd.DataFrame) -> pd.DataFrame:
    """Select top 20 positive distinctive terms."""
    return df.head(20)[["token", "difference_per_1000"]].copy()


def transform_distinctive_terms_negative_vs_positive(df: pd.DataFrame) -> pd.DataFrame:
    """Select top 20 negative distinctive terms."""
    return df.head(20)[["token", "difference_per_1000"]].copy()


def transform_emotion_distribution_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Return emotion distribution metric for charting."""
    return df[["emotion", "mean_per_100_tokens"]].copy()


def transform_emotion_by_polarity(df: pd.DataFrame) -> pd.DataFrame:
    """Reshape emotion-by-polarity into chart-ready wide table."""
    temp = df.copy()
    temp = temp[temp["segment"].isin(["positive", "negative"])].copy()

    rows = []
    emotions = [
        "positive", "joy", "anticipation", "trust", "surprise",
        "negative", "fear", "sadness", "anger", "disgust",
    ]

    for emotion in emotions:
        row = {"emotion": emotion}
        for segment in ["positive", "negative"]:
            subset = temp[temp["segment"] == segment]
            value_col = f"{emotion}_mean_per_100_tokens"
            row[segment] = (
                pd.to_numeric(subset.iloc[0][value_col], errors="coerce")
                if not subset.empty and value_col in subset.columns
                else None
            )
        rows.append(row)

    return pd.DataFrame(rows)


def transform_emotion_by_month_selected(df: pd.DataFrame) -> pd.DataFrame:
    """Return selected emotion trajectories."""
    columns = [
        "review_created_year_month",
        "joy_mean_per_100_tokens",
        "trust_mean_per_100_tokens",
        "negative_mean_per_100_tokens",
        "anger_mean_per_100_tokens",
    ]
    available = [col for col in columns if col in df.columns]
    return df[available].copy()


def transform_theme_distribution_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Return theme distribution metric for charting."""
    return df[["theme", "review_share_with_theme"]].copy()


def transform_theme_by_polarity(df: pd.DataFrame) -> pd.DataFrame:
    """Reshape theme-by-polarity into chart-ready wide table."""
    temp = df.copy()
    temp = temp[temp["segment"].isin(["positive", "negative"])].copy()

    theme_prefixes = []
    for column in temp.columns:
        if column.endswith("_review_share"):
            prefix = column[:-len("_review_share")]
            if prefix not in ("segment",):
                theme_prefixes.append(prefix)

    theme_prefixes = sorted(set(theme_prefixes))

    rows = []
    for theme in theme_prefixes:
        row = {"theme": theme}
        for segment in ["positive", "negative"]:
            subset = temp[temp["segment"] == segment]
            value_col = f"{theme}_review_share"
            row[segment] = (
                pd.to_numeric(subset.iloc[0][value_col], errors="coerce")
                if not subset.empty and value_col in subset.columns
                else None
            )
        rows.append(row)

    return pd.DataFrame(rows)


def transform_theme_by_month_selected(df: pd.DataFrame) -> pd.DataFrame:
    """Return selected theme trajectories."""
    columns = [
        "review_created_year_month",
        "difficulty_review_share",
        "combat_review_share",
        "performance_technical_review_share",
        "story_review_share",
    ]
    available = [col for col in columns if col in df.columns]
    return df[available].copy()


TRANSFORMS: Dict[str, Callable[[pd.DataFrame], pd.DataFrame]] = {
    "polarity_distribution": transform_polarity_distribution,
    "language_distribution_top10": transform_language_distribution_top10,
    "polarity_by_playtime_band": transform_polarity_by_playtime_band,
    "monthly_review_volume": transform_monthly_review_volume,
    "monthly_polarity_trends": transform_monthly_polarity_trends,
    "monthly_positive_percentage": transform_monthly_positive_percentage,
    "monthly_language_trends_top5": transform_monthly_language_trends_top5,
    "developer_response_timeline": transform_developer_response_timeline,
    "top_unigrams_positive": transform_top_unigrams_positive,
    "top_unigrams_negative": transform_top_unigrams_negative,
    "distinctive_terms_positive_vs_negative": transform_distinctive_terms_positive_vs_negative,
    "distinctive_terms_negative_vs_positive": transform_distinctive_terms_negative_vs_positive,
    "emotion_distribution_summary": transform_emotion_distribution_summary,
    "emotion_by_polarity": transform_emotion_by_polarity,
    "emotion_by_month_selected": transform_emotion_by_month_selected,
    "theme_distribution_summary": transform_theme_distribution_summary,
    "theme_by_polarity": transform_theme_by_polarity,
    "theme_by_month_selected": transform_theme_by_month_selected,
}


# ============================================================================
# EXCEL HELPERS
# ============================================================================

def get_family_output_dir(paths: Dict[str, Path], family: str) -> Path:
    """Map figure family to output directory."""
    mapping = {
        "basic_metrics": paths["basic_figures_dir"],
        "temporal_metrics": paths["temporal_figures_dir"],
        "text_metrics": paths["text_figures_dir"],
        "emotion_metrics": paths["emotion_figures_dir"],
        "theme_metrics": paths["theme_figures_dir"],
    }
    if family not in mapping:
        raise ValueError(f"Unknown figure family: {family}")
    return mapping[family]


def write_dataframe_to_sheet(ws, df: pd.DataFrame) -> None:
    """Write a dataframe to an openpyxl worksheet."""
    if df.empty:
        ws["A1"] = "No data available"
        return

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(column))
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=normalise_excel_value(value))


def normalise_excel_value(value: Any) -> Any:
    """Convert pandas values into Excel-compatible scalar values."""
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            value = value.item()
        except Exception:
            return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    if isinstance(value, (int, str, bool)):
        return value
    return str(value)


def create_chart_from_sheet(
    wb: Workbook,
    chart_sheet_name: str,
    data_sheet_name: str,
    df: pd.DataFrame,
    chart_type: str,
    title: str,
) -> None:
    """Create an Excel chart from a data sheet."""
    chart_ws = wb.create_sheet(chart_sheet_name)
    chart_ws["A1"] = title
    chart_ws["A1"].font = Font(bold=True, size=14)

    if df.empty or len(df.columns) < 2:
        chart_ws["A3"] = "No chart could be created because the source table is empty or incomplete."
        return

    min_row = 1
    max_row = len(df) + 1
    min_col = 2
    max_col = len(df.columns)

    category_ref = Reference(
        wb[data_sheet_name],
        min_col=1,
        min_row=2,
        max_row=max_row,
    )
    data_ref = Reference(
        wb[data_sheet_name],
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )

    if chart_type == "line":
        chart = LineChart()
        chart.style = 2
        chart.y_axis.title = "Value"
        chart.x_axis.title = str(df.columns[0])

    elif chart_type in {"bar", "bar_stacked"}:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 2
        chart.y_axis.title = str(df.columns[0])
        chart.x_axis.title = "Value"
        if chart_type == "bar_stacked":
            chart.grouping = "stacked"
            chart.overlap = 100
    else:
        raise ValueError(f"Unsupported chart type: {chart_type}")

    chart.title = title
    chart.height = 12
    chart.width = 22
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(category_ref)
    chart.legend.position = "r"

    chart_ws.add_chart(chart, "A3")


def create_metadata_sheet(
    wb: Workbook,
    spec: FigureSpec,
    source_path: Path,
    output_path: Path,
) -> None:
    """Create metadata sheet for one figure workbook."""
    ws = wb.create_sheet("metadata")
    metadata_rows = [
        ("figure_id", spec.figure_id),
        ("filename", spec.filename),
        ("family", spec.family),
        ("title", spec.title),
        ("chart_type", spec.chart_type),
        ("source_table", str(source_path)),
        ("output_file", str(output_path)),
        ("transform_name", spec.transform_name),
        ("category_column", spec.category_column),
        ("value_columns", ", ".join(spec.value_columns) if spec.value_columns else ""),
        ("notes", spec.notes),
    ]

    for row_idx, (key, value) in enumerate(metadata_rows, start=1):
        ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)


# ============================================================================
# FIGURE WORKBOOK CREATION
# ============================================================================

def create_figure_workbook(
    spec: FigureSpec,
    paths: Dict[str, Path],
) -> Dict[str, Any]:
    """Create one Excel workbook for one figure spec."""
    source_path = resolve_source_table_path(paths, spec.source_table)
    source_df = load_source_table(source_path)

    if spec.transform_name not in TRANSFORMS:
        raise ValueError(f"Unknown transform: {spec.transform_name}")

    transformed_df = TRANSFORMS[spec.transform_name](source_df)

    output_dir = get_family_output_dir(paths, spec.family)
    output_path = output_dir / spec.filename

    wb = Workbook()
    default_ws = wb.active
    default_ws.title = "data"

    write_dataframe_to_sheet(default_ws, transformed_df)
    create_chart_from_sheet(
        wb=wb,
        chart_sheet_name="chart",
        data_sheet_name="data",
        df=transformed_df,
        chart_type=spec.chart_type,
        title=spec.title,
    )
    create_metadata_sheet(
        wb=wb,
        spec=spec,
        source_path=source_path,
        output_path=output_path,
    )

    wb.save(output_path)

    return {
        "figure_id": spec.figure_id,
        "filename": spec.filename,
        "family": spec.family,
        "title": spec.title,
        "chart_type": spec.chart_type,
        "source_table": str(source_path),
        "output_file": str(output_path),
        "row_count": int(len(transformed_df)),
        "column_count": int(len(transformed_df.columns)),
        "notes": spec.notes,
    }


# ============================================================================
# SERIALISATION HELPERS
# ============================================================================

def to_serialisable(value: Any) -> Any:
    """Convert values to JSON-safe Python values."""
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            value = value.item()
        except Exception:
            return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    if isinstance(value, (int, str, bool)):
        return value
    return str(value)


# ============================================================================
# MAIN
# ============================================================================

def main() -> None:
    """Main entry point."""
    args = parse_arguments()

    config_path = Path(args.config).resolve()
    config = load_game_config(config_path)

    repository_root = get_repository_root()
    paths = build_paths(repository_root, config)
    ensure_directories(paths)
    configure_logging(paths["analysis_log"])

    print(f"[INFO] Loading configuration: {config_path}")
    print(f"[INFO] Repository root: {repository_root}")
    print(f"[INFO] Game slug: {config.game_slug}")
    print(f"[INFO] Figures root: {paths['figures_root']}")
    print(f"[INFO] Figure index CSV: {paths['figure_index_csv']}")
    print(f"[INFO] Summary JSON: {paths['summary_json']}")

    logging.info("Starting Excel figure generation.")
    logging.info("Configuration file: %s", config_path)
    logging.info("Game title: %s", config.game_title)
    logging.info("App ID: %s", config.app_id)
    logging.info("Game slug: %s", config.game_slug)

    specs = get_figure_specs(config)
    logging.info("Figure specs loaded: %s", len(specs))

    figure_rows: List[Dict[str, Any]] = []

    for spec in specs:
        logging.info("Creating figure workbook: %s", spec.filename)
        result = create_figure_workbook(spec, paths)
        figure_rows.append(result)

    figure_index_df = pd.DataFrame(figure_rows)
    figure_index_df.to_csv(paths["figure_index_csv"], index=False, encoding="utf-8-sig")

    summary_payload = {
        "app_id": config.app_id,
        "game_slug": config.game_slug,
        "game_title": config.game_title,
        "generated_at_utc": pd.Timestamp.utcnow().isoformat(),
        "figure_count": len(figure_rows),
        "figures": [
            {key: to_serialisable(value) for key, value in row.items()}
            for row in figure_rows
        ],
    }

    with paths["summary_json"].open("w", encoding="utf-8") as handle:
        json.dump(summary_payload, handle, ensure_ascii=False, indent=2)

    logging.info("Figure index CSV written to: %s", paths["figure_index_csv"])
    logging.info("Summary JSON written to: %s", paths["summary_json"])
    logging.info("Excel figure generation complete.")
    print("[INFO] Excel figure generation complete.")


if __name__ == "__main__":
    main()