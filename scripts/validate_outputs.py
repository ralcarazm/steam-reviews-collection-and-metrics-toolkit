#!/usr/bin/env python3
"""
validate_outputs.py

Validate the outputs generated for one selected game configuration.

This script:
- reads a per-game JSON configuration file;
- checks the existence of key outputs across raw, processed, tables, and figures;
- validates basic structural expectations such as required columns;
- validates selected consistency rules across outputs;
- writes a validation summary JSON file;
- writes a validation log.

Usage examples:

    python scripts/validate_outputs.py --config config/games/game_02.json

Author:
    Rubén Alcaraz Martínez

Licence:
    GNU General Public License v3.0
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


# ============================================================================
# DATA MODELS
# ============================================================================

@dataclass
class GameConfig:
    """Configuration values required for validation."""
    app_id: int
    game_slug: str
    game_title: str


# ============================================================================
# ARGUMENT PARSING
# ============================================================================

def parse_arguments() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Validate repository outputs for one game."
    )
    parser.add_argument(
        "--config",
        required=True,
        help="Path to the JSON configuration file for the target game.",
    )
    return parser.parse_args()


# ============================================================================
# CONFIGURATION LOADING
# ============================================================================

def load_game_config(config_path: Path) -> GameConfig:
    """Load and validate the game configuration file."""
    if not config_path.exists():
        raise FileNotFoundError(f"Configuration file not found: {config_path}")

    with config_path.open("r", encoding="utf-8") as handle:
        raw_config = json.load(handle)

    required_fields = ["app_id", "game_slug", "game_title"]
    missing_fields = [field for field in required_fields if field not in raw_config]
    if missing_fields:
        raise ValueError(
            f"Missing required configuration fields: {', '.join(missing_fields)}"
        )

    return GameConfig(
        app_id=int(raw_config["app_id"]),
        game_slug=str(raw_config["game_slug"]).strip(),
        game_title=str(raw_config["game_title"]).strip(),
    )


# ============================================================================
# PATH MANAGEMENT
# ============================================================================

def get_repository_root() -> Path:
    """
    Resolve the repository root assuming this file lives at:
    scripts/validate_outputs.py
    """
    return Path(__file__).resolve().parents[1]


def build_paths(repository_root: Path, config: GameConfig) -> Dict[str, Path]:
    """Build key repository paths for validation."""
    raw_root = repository_root / "data" / "raw" / "steam_reviews" / config.game_slug
    processed_root = repository_root / "data" / "processed" / "steam_reviews" / config.game_slug
    results_root = repository_root / "results" / config.game_slug
    metrics_dir = processed_root / "metrics"

    return {
        "repository_root": repository_root,
        "raw_root": raw_root,
        "processed_root": processed_root,
        "results_root": results_root,
        "metrics_dir": metrics_dir,
        "validation_log": metrics_dir / f"{config.game_slug}_validation.log",
        "validation_summary_json": metrics_dir / f"{config.game_slug}_validation_summary.json",
        "raw_combined_csv": raw_root / "combined" / f"{config.game_slug}_reviews_all.csv",
        "cleaned_csv": processed_root / "cleaned" / f"{config.game_slug}_reviews_cleaned.csv",
        "figure_index_csv": results_root / "figures" / f"{config.game_slug}_figure_index.csv",
        "wordcloud_dir": results_root / "figures" / "word_cloud",
    }


# ============================================================================
# LOGGING
# ============================================================================

def configure_logging(log_file: Path) -> None:
    """Configure console and file logging."""
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )


# ============================================================================
# HELPERS
# ============================================================================

def utc_now_iso() -> str:
    """Return current UTC timestamp as ISO 8601 string."""
    return datetime.now(timezone.utc).isoformat()


def safe_read_csv(path: Path) -> pd.DataFrame:
    """Read a CSV safely."""
    return pd.read_csv(path, encoding="utf-8-sig", low_memory=False)


def make_check(
    check_id: str,
    passed: bool,
    severity: str,
    message: str,
    details: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Create a validation check record."""
    return {
        "check_id": check_id,
        "passed": passed,
        "severity": severity,
        "message": message,
        "details": details or {},
    }


def file_exists_check(path: Path, check_id: str, severity: str = "error") -> Dict[str, Any]:
    """Create a file-existence check."""
    exists = path.exists()
    return make_check(
        check_id=check_id,
        passed=exists,
        severity=severity,
        message=f"{'Found' if exists else 'Missing'}: {path}",
        details={"path": str(path)},
    )


def required_columns_check(
    df: pd.DataFrame,
    required_columns: List[str],
    check_id: str,
    table_path: Path,
    severity: str = "error",
) -> Dict[str, Any]:
    """Validate required columns."""
    missing = [col for col in required_columns if col not in df.columns]
    return make_check(
        check_id=check_id,
        passed=not missing,
        severity=severity,
        message=(
            f"Required columns present in {table_path.name}"
            if not missing
            else f"Missing required columns in {table_path.name}: {', '.join(missing)}"
        ),
        details={
            "path": str(table_path),
            "required_columns": required_columns,
            "missing_columns": missing,
        },
    )


def worksheet_presence_check(
    workbook_path: Path,
    required_names: List[str],
    check_id: str,
    severity: str = "error",
) -> Dict[str, Any]:
    """Validate worksheet presence using openpyxl."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return make_check(
            check_id=check_id,
            passed=False,
            severity=severity,
            message=f"Could not import openpyxl while checking workbook sheets: {exc}",
            details={"path": str(workbook_path)},
        )

    try:
        wb = load_workbook(workbook_path, read_only=True)
        sheet_names = wb.sheetnames
    except Exception as exc:
        return make_check(
            check_id=check_id,
            passed=False,
            severity=severity,
            message=f"Could not open workbook: {workbook_path.name} | {exc}",
            details={"path": str(workbook_path)},
        )

    missing = [name for name in required_names if name not in sheet_names]
    return make_check(
        check_id=check_id,
        passed=not missing,
        severity=severity,
        message=(
            f"Required worksheets present in {workbook_path.name}"
            if not missing
            else f"Missing worksheets in {workbook_path.name}: {', '.join(missing)}"
        ),
        details={
            "path": str(workbook_path),
            "sheet_names": sheet_names,
            "missing_worksheets": missing,
        },
    )


# ============================================================================
# VALIDATION RULES
# ============================================================================

def validate_required_files(paths: Dict[str, Path], config: GameConfig) -> List[Dict[str, Any]]:
    """Validate required key files."""
    game_slug = config.game_slug
    processed_metrics = paths["processed_root"] / "metrics"
    results_tables = paths["results_root"] / "tables"

    checks = [
        file_exists_check(paths["raw_combined_csv"], "required_raw_combined_csv"),
        file_exists_check(paths["cleaned_csv"], "required_cleaned_csv"),
        file_exists_check(processed_metrics / f"{game_slug}_basic_metrics_summary.json", "required_basic_metrics_summary"),
        file_exists_check(processed_metrics / f"{game_slug}_temporal_metrics_summary.json", "required_temporal_metrics_summary"),
        file_exists_check(processed_metrics / f"{game_slug}_text_metrics_summary.json", "required_text_metrics_summary"),
        file_exists_check(processed_metrics / f"{game_slug}_emotion_metrics_summary.json", "required_emotion_metrics_summary"),
        file_exists_check(processed_metrics / f"{game_slug}_theme_metrics_summary.json", "required_theme_metrics_summary"),
        file_exists_check(processed_metrics / f"{game_slug}_figures_summary.json", "required_figures_summary"),
        file_exists_check(processed_metrics / f"{game_slug}_word_cloud_summary.json", "required_wordcloud_summary"),
        file_exists_check(paths["figure_index_csv"], "required_figure_index_csv"),
        file_exists_check(results_tables / "basic_metrics" / f"{game_slug}_polarity_summary.csv", "required_polarity_summary_csv"),
        file_exists_check(results_tables / "temporal_metrics" / f"{game_slug}_monthly_polarity_trends.csv", "required_monthly_polarity_trends_csv"),
        file_exists_check(results_tables / "emotion_metrics" / f"{game_slug}_emotion_by_polarity.csv", "required_emotion_by_polarity_csv"),
        file_exists_check(results_tables / "theme_metrics" / f"{game_slug}_theme_by_polarity.csv", "required_theme_by_polarity_csv"),
    ]
    return checks


def validate_cleaned_csv(paths: Dict[str, Path]) -> List[Dict[str, Any]]:
    """Validate cleaned CSV structure."""
    checks: List[Dict[str, Any]] = []
    path = paths["cleaned_csv"]

    if not path.exists():
        return [file_exists_check(path, "cleaned_csv_exists")]

    df = safe_read_csv(path)

    checks.append(
        required_columns_check(
            df=df,
            required_columns=["recommendationid", "review", "language", "voted_up"],
            check_id="cleaned_csv_required_columns",
            table_path=path,
        )
    )

    checks.append(
        make_check(
            check_id="cleaned_csv_non_empty",
            passed=len(df) > 0,
            severity="error",
            message=f"{path.name} {'contains rows' if len(df) > 0 else 'is empty'}",
            details={"row_count": int(len(df)), "path": str(path)},
        )
    )

    if "recommendationid" in df.columns:
        duplicated = int(df["recommendationid"].astype("string").fillna("").duplicated().sum())
        checks.append(
            make_check(
                check_id="cleaned_csv_recommendationid_uniqueness",
                passed=duplicated == 0,
                severity="warning",
                message=(
                    "No duplicated recommendation IDs in cleaned CSV"
                    if duplicated == 0
                    else f"Found duplicated recommendation IDs in cleaned CSV: {duplicated}"
                ),
                details={"duplicated_count": duplicated, "path": str(path)},
            )
        )

    return checks


def validate_basic_metrics(paths: Dict[str, Path], config: GameConfig) -> List[Dict[str, Any]]:
    """Validate basic metrics outputs and selected consistency rules."""
    checks: List[Dict[str, Any]] = []
    polarity_path = paths["results_root"] / "tables" / "basic_metrics" / f"{config.game_slug}_polarity_summary.csv"

    if not polarity_path.exists():
        return [file_exists_check(polarity_path, "basic_polarity_summary_exists")]

    df = safe_read_csv(polarity_path)
    checks.append(
        required_columns_check(
            df=df,
            required_columns=["metric", "value"],
            check_id="basic_polarity_summary_required_columns",
            table_path=polarity_path,
        )
    )

    metrics = {}
    try:
        metrics = dict(zip(df["metric"].astype(str), pd.to_numeric(df["value"], errors="coerce")))
    except Exception:
        pass

    total_reviews = metrics.get("total_reviews")
    positive_reviews = metrics.get("positive_reviews")
    negative_reviews = metrics.get("negative_reviews")

    if total_reviews is not None and positive_reviews is not None and negative_reviews is not None:
        passed = int(total_reviews) == int(positive_reviews + negative_reviews)
        checks.append(
            make_check(
                check_id="basic_polarity_total_consistency",
                passed=passed,
                severity="warning",
                message=(
                    "Basic polarity totals are consistent"
                    if passed
                    else "Basic polarity totals are inconsistent"
                ),
                details={
                    "total_reviews": None if pd.isna(total_reviews) else int(total_reviews),
                    "positive_reviews": None if pd.isna(positive_reviews) else int(positive_reviews),
                    "negative_reviews": None if pd.isna(negative_reviews) else int(negative_reviews),
                },
            )
        )

    return checks


def validate_temporal_metrics(paths: Dict[str, Path], config: GameConfig) -> List[Dict[str, Any]]:
    """Validate temporal metrics structure."""
    checks: List[Dict[str, Any]] = []
    month_path = paths["results_root"] / "tables" / "temporal_metrics" / f"{config.game_slug}_monthly_polarity_trends.csv"

    if not month_path.exists():
        return [file_exists_check(month_path, "temporal_monthly_polarity_trends_exists")]

    df = safe_read_csv(month_path)

    checks.append(
        required_columns_check(
            df=df,
            required_columns=[
                "review_created_date_month",
                "review_count",
                "positive_reviews",
                "negative_reviews",
                "positive_percentage",
            ],
            check_id="temporal_monthly_polarity_trends_required_columns",
            table_path=month_path,
        )
    )

    checks.append(
        make_check(
            check_id="temporal_monthly_polarity_trends_non_empty",
            passed=len(df) > 0,
            severity="warning",
            message=(
                "Temporal monthly polarity trends contain rows"
                if len(df) > 0
                else "Temporal monthly polarity trends are empty"
            ),
            details={"row_count": int(len(df)), "path": str(month_path)},
        )
    )

    return checks


def validate_text_metrics(paths: Dict[str, Path]) -> List[Dict[str, Any]]:
    """Validate text metrics structure."""
    checks: List[Dict[str, Any]] = []
    global_dir = paths["results_root"] / "tables" / "text_metrics" / "global"
    pos_path = global_dir / "top_unigrams_positive.csv"
    neg_path = global_dir / "top_unigrams_negative.csv"

    for check_id, path in [
        ("text_top_unigrams_positive_exists", pos_path),
        ("text_top_unigrams_negative_exists", neg_path),
    ]:
        checks.append(file_exists_check(path, check_id))

    for check_id, path in [
        ("text_top_unigrams_positive_required_columns", pos_path),
        ("text_top_unigrams_negative_required_columns", neg_path),
    ]:
        if path.exists():
            df = safe_read_csv(path)
            checks.append(
                required_columns_check(
                    df=df,
                    required_columns=["rank", "token", "frequency"],
                    check_id=check_id,
                    table_path=path,
                )
            )

    return checks


def validate_emotion_metrics(paths: Dict[str, Path], config: GameConfig) -> List[Dict[str, Any]]:
    """Validate emotion metrics structure."""
    checks: List[Dict[str, Any]] = []
    path = paths["results_root"] / "tables" / "emotion_metrics" / f"{config.game_slug}_emotion_by_polarity.csv"

    if not path.exists():
        return [file_exists_check(path, "emotion_by_polarity_exists")]

    df = safe_read_csv(path)

    checks.append(
        required_columns_check(
            df=df,
            required_columns=["segment", "review_count", "total_tokens", "emotion_token_count"],
            check_id="emotion_by_polarity_required_columns",
            table_path=path,
        )
    )

    expected_segments = {"overall", "positive", "negative"}
    present_segments = set(df["segment"].astype(str).tolist()) if "segment" in df.columns else set()
    missing_segments = sorted(expected_segments - present_segments)

    checks.append(
        make_check(
            check_id="emotion_by_polarity_expected_segments",
            passed=not missing_segments,
            severity="warning",
            message=(
                "Emotion by polarity contains expected segments"
                if not missing_segments
                else f"Emotion by polarity is missing segments: {', '.join(missing_segments)}"
            ),
            details={"present_segments": sorted(present_segments), "missing_segments": missing_segments},
        )
    )

    return checks


def validate_theme_metrics(paths: Dict[str, Path], config: GameConfig) -> List[Dict[str, Any]]:
    """Validate theme metrics structure."""
    checks: List[Dict[str, Any]] = []
    path = paths["results_root"] / "tables" / "theme_metrics" / f"{config.game_slug}_theme_by_polarity.csv"

    if not path.exists():
        return [file_exists_check(path, "theme_by_polarity_exists")]

    df = safe_read_csv(path)

    checks.append(
        required_columns_check(
            df=df,
            required_columns=["segment", "review_count", "total_tokens", "theme_match_count_total"],
            check_id="theme_by_polarity_required_columns",
            table_path=path,
        )
    )

    expected_segments = {"overall", "positive", "negative"}
    present_segments = set(df["segment"].astype(str).tolist()) if "segment" in df.columns else set()
    missing_segments = sorted(expected_segments - present_segments)

    checks.append(
        make_check(
            check_id="theme_by_polarity_expected_segments",
            passed=not missing_segments,
            severity="warning",
            message=(
                "Theme by polarity contains expected segments"
                if not missing_segments
                else f"Theme by polarity is missing segments: {', '.join(missing_segments)}"
            ),
            details={"present_segments": sorted(present_segments), "missing_segments": missing_segments},
        )
    )

    return checks


def validate_figures(paths: Dict[str, Path], config: GameConfig) -> List[Dict[str, Any]]:
    """Validate figure outputs and workbook structure."""
    checks: List[Dict[str, Any]] = []
    figure_index_path = paths["figure_index_csv"]

    if not figure_index_path.exists():
        return [file_exists_check(figure_index_path, "figure_index_exists")]

    figure_index_df = safe_read_csv(figure_index_path)
    checks.append(
        required_columns_check(
            df=figure_index_df,
            required_columns=[
                "figure_id",
                "filename",
                "family",
                "title",
                "chart_type",
                "source_table",
                "output_file",
                "row_count",
                "column_count",
                "notes",
            ],
            check_id="figure_index_required_columns",
            table_path=figure_index_path,
        )
    )

    checks.append(
        make_check(
            check_id="figure_index_non_empty",
            passed=len(figure_index_df) > 0,
            severity="warning",
            message=(
                "Figure index contains rows"
                if len(figure_index_df) > 0
                else "Figure index is empty"
            ),
            details={"row_count": int(len(figure_index_df)), "path": str(figure_index_path)},
        )
    )

    workbook_paths: List[Path] = []
    if "output_file" in figure_index_df.columns:
        workbook_paths = [Path(str(value)) for value in figure_index_df["output_file"].dropna().tolist()]

    existing_workbooks = [path for path in workbook_paths if path.exists()]
    missing_workbooks = [str(path) for path in workbook_paths if not path.exists()]

    checks.append(
        make_check(
            check_id="figure_workbooks_exist",
            passed=not missing_workbooks,
            severity="warning",
            message=(
                "All figure workbooks listed in the figure index exist"
                if not missing_workbooks
                else "Some figure workbooks listed in the figure index are missing"
            ),
            details={
                "existing_count": len(existing_workbooks),
                "missing_count": len(missing_workbooks),
                "missing_workbooks": missing_workbooks,
            },
        )
    )

    for idx, workbook_path in enumerate(existing_workbooks, start=1):
        checks.append(
            worksheet_presence_check(
                workbook_path=workbook_path,
                required_names=["data", "chart", "metadata"],
                check_id=f"figure_workbook_required_sheets_{idx}",
                severity="warning",
            )
        )

    return checks


def validate_wordcloud_outputs(paths: Dict[str, Path], config: GameConfig) -> List[Dict[str, Any]]:
    """Validate word-cloud outputs."""
    checks: List[Dict[str, Any]] = []
    game_slug = config.game_slug
    wordcloud_dir = paths["wordcloud_dir"]

    expected_files = [
        wordcloud_dir / f"{game_slug}_wordcloud_unigrams.png",
        wordcloud_dir / f"{game_slug}_wordcloud_bigrams.png",
        wordcloud_dir / f"{game_slug}_wordcloud_trigrams.png",
        wordcloud_dir / f"{game_slug}_wordcloud_unigrams_frequencies.csv",
        wordcloud_dir / f"{game_slug}_wordcloud_bigrams_frequencies.csv",
        wordcloud_dir / f"{game_slug}_wordcloud_trigrams_frequencies.csv",
    ]

    for idx, path in enumerate(expected_files, start=1):
        checks.append(file_exists_check(path, f"wordcloud_expected_file_{idx}", severity="warning"))

    unigram_csv = wordcloud_dir / f"{game_slug}_wordcloud_unigrams_frequencies.csv"
    if unigram_csv.exists():
        df = safe_read_csv(unigram_csv)
        checks.append(
            required_columns_check(
                df=df,
                required_columns=["rank", "unigram", "frequency"],
                check_id="wordcloud_unigram_frequency_columns",
                table_path=unigram_csv,
                severity="warning",
            )
        )

    bigram_csv = wordcloud_dir / f"{game_slug}_wordcloud_bigrams_frequencies.csv"
    if bigram_csv.exists():
        df = safe_read_csv(bigram_csv)
        checks.append(
            required_columns_check(
                df=df,
                required_columns=["rank", "bigram", "frequency"],
                check_id="wordcloud_bigram_frequency_columns",
                table_path=bigram_csv,
                severity="warning",
            )
        )

    trigram_csv = wordcloud_dir / f"{game_slug}_wordcloud_trigrams_frequencies.csv"
    if trigram_csv.exists():
        df = safe_read_csv(trigram_csv)
        checks.append(
            required_columns_check(
                df=df,
                required_columns=["rank", "trigram", "frequency"],
                check_id="wordcloud_trigram_frequency_columns",
                table_path=trigram_csv,
                severity="warning",
            )
        )

    return checks


# ============================================================================
# MAIN
# ============================================================================

def main() -> None:
    """Main entry point."""
    args = parse_arguments()

    config_path = Path(args.config).resolve()
    config = load_game_config(config_path)

    repository_root = get_repository_root()
    paths = build_paths(repository_root, config)
    configure_logging(paths["validation_log"])

    print(f"[INFO] Loading configuration: {config_path}")
    print(f"[INFO] Repository root: {repository_root}")
    print(f"[INFO] Game slug: {config.game_slug}")
    print(f"[INFO] Validation log: {paths['validation_log']}")
    print(f"[INFO] Validation summary: {paths['validation_summary_json']}")

    logging.info("Starting output validation.")
    logging.info("Game title: %s", config.game_title)
    logging.info("App ID: %s", config.app_id)
    logging.info("Game slug: %s", config.game_slug)

    checks: List[Dict[str, Any]] = []
    checks.extend(validate_required_files(paths, config))
    checks.extend(validate_cleaned_csv(paths))
    checks.extend(validate_basic_metrics(paths, config))
    checks.extend(validate_temporal_metrics(paths, config))
    checks.extend(validate_text_metrics(paths))
    checks.extend(validate_emotion_metrics(paths, config))
    checks.extend(validate_theme_metrics(paths, config))
    checks.extend(validate_figures(paths, config))
    checks.extend(validate_wordcloud_outputs(paths, config))

    passed_count = sum(1 for check in checks if check["passed"])
    failed_count = sum(1 for check in checks if not check["passed"])
    error_failures = sum(1 for check in checks if not check["passed"] and check["severity"] == "error")
    warning_failures = sum(1 for check in checks if not check["passed"] and check["severity"] == "warning")

    summary = {
        "app_id": config.app_id,
        "game_slug": config.game_slug,
        "game_title": config.game_title,
        "generated_at_utc": utc_now_iso(),
        "check_count_total": len(checks),
        "check_count_passed": passed_count,
        "check_count_failed": failed_count,
        "error_failures": error_failures,
        "warning_failures": warning_failures,
        "final_status": "failed" if error_failures > 0 else "passed",
        "checks": checks,
    }

    with paths["validation_summary_json"].open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)

    logging.info("Validation summary written to: %s", paths["validation_summary_json"])
    logging.info("Validation complete. final_status=%s", summary["final_status"])

    if summary["final_status"] == "failed":
        print("[INFO] Validation finished with errors.")
        sys.exit(1)

    print("[INFO] Validation finished successfully.")


if __name__ == "__main__":
    main()